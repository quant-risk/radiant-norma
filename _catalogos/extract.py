"""Extração refinada v3 — detecta header em qualquer linha, código em qualquer coluna."""

import json
import re
from pathlib import Path
from datetime import datetime

import xlrd
import openpyxl

ROOT = Path("/Users/henrique/Downloads/cadocs")
OUT = ROOT / "_catalogos"
OUT.mkdir(exist_ok=True)

LOG = []


def log(msg):
    ts = datetime.now().isoformat(timespec="seconds")
    LOG.append(f"[{ts}] {msg}")
    print(msg)


def read_xls(path: Path):
    try:
        book = xlrd.open_workbook(str(path), formatting_info=False)
        sheets = []
        for s_idx in range(len(book.sheet_names())):
            sh = book.sheet_by_index(s_idx)
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    val = sh.cell_value(r, c)
                    if isinstance(val, float) and val.is_integer():
                        val = int(val)
                    row.append(val if val != "" else None)
                if any(c is not None and str(c).strip() != "" for c in row):
                    rows.append(row)
            sheets.append({"name": book.sheet_names()[s_idx], "rows": rows})
        return sheets
    except Exception as e:
        log(f"  ! Erro XLS {path.name}: {e}")
        return []


def read_xlsx(path: Path):
    try:
        book = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheets = []
        for s_name in book.sheetnames:
            sh = book[s_name]
            rows = []
            for row in sh.iter_rows(values_only=True):
                row_clean = [v if (v is not None and not (isinstance(v, str) and v.strip() == "")) else None for v in row]
                if any(c is not None for c in row_clean):
                    rows.append(row_clean)
            sheets.append({"name": s_name, "rows": rows})
        book.close()
        return sheets
    except Exception as e:
        log(f"  ! Erro XLSX {path.name}: {e}")
        return []


def read_any(path: Path):
    return read_xls(path) if path.suffix.lower() == ".xls" else read_xlsx(path)


def detect_header_row(rows: list, max_search: int = 10) -> int:
    """Encontra linha de cabeçalho em qualquer das primeiras N linhas."""
    keywords = ["cód", "codigo", "chave", "código", "code", "critica", "crití"]
    for i in range(min(max_search, len(rows))):
        row_str = " ".join(str(c or "").strip().lower() for c in rows[i])
        if any(k in row_str for k in keywords):
            return i
    return -1


def detect_code_column(header: list) -> int:
    """Detecta coluna onde está o código."""
    header_lower = [str(c or "").strip().lower() for c in header]
    # Procura coluna com nome contendo "cód", "codigo", "chave", "code"
    for i, h in enumerate(header_lower):
        if any(k in h for k in ("cód", "codigo", "chave", "code", "crití")):
            return i
    return 0


def extract_criticas_from_sheet(sheet_rows: list, sheet_name: str, cadoc_code: str) -> list:
    """Extrai críticas de uma sheet."""
    criticas = []
    header_idx = detect_header_row(sheet_rows)
    if header_idx == -1:
        return criticas
    header = [str(c or "").strip().lower() for c in sheet_rows[header_idx]]
    codigo_col = detect_code_column(header)

    # Detecta coluna de data-base inicio
    data_inicio_col = -1
    for i, h in enumerate(header):
        if "data" in h and "inic" in h:
            data_inicio_col = i
            break
        if h in ("vigência inicial", "vigencia inicial", "data-base início", "data-base inicio"):
            data_inicio_col = i
            break

    for row in sheet_rows[header_idx + 1:]:
        if not row or codigo_col >= len(row) or row[codigo_col] is None:
            continue
        codigo_raw = row[codigo_col]
        codigo = str(codigo_raw).strip()
        # Padrões: B01, B02, 2001, 4678, ELIM0001, etc.
        if not re.match(r"^[A-Z]{0,5}\d{2,5}", codigo):
            continue
        record = {
            "cadoc": cadoc_code,
            "sheet": sheet_name,
            "codigo": codigo,
        }
        for i, h in enumerate(header):
            if i == codigo_col or not h:
                continue
            if i < len(row) and row[i] is not None:
                val = row[i]
                if isinstance(val, str):
                    val = val.strip()
                # Normaliza data-base inicio
                if i == data_inicio_col:
                    val = normalize_date(val)
                record[h] = val
        if len(record) > 3:  # tem pelo menos cadoc + sheet + codigo + 1 campo
            criticas.append(record)
    return criticas


def normalize_date(val):
    """Normaliza data para ISO 8601 (YYYY-MM-DD).

    Suporta:
    - int 201501 → "2015-01-01"
    - float Excel serial (43862.0) → "2020-02-05"
    - "AAAA-MM" → "AAAA-MM-01"
    - "AAAA-MM-DD" → mantém
    - datetime.date → ISO
    - None → None
    """
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("nan", "none", "null", "n/a"):
            return None
        # AAAAMM (6 dígitos sem hífen) - mas int seria mais comum
        if re.match(r"^\d{6}$", s):
            return f"{s[:4]}-{s[4:6]}-01"
        # AAAA-MM
        if re.match(r"^\d{4}-\d{2}$", s):
            return f"{s}-01"
        # AAAA-MM-DD
        if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
            return s
        return s  # formato desconhecido, mantém
    if isinstance(val, int):
        # Provavelmente AAAAMM
        if 190001 <= val <= 210012:
            s = str(val)
            if len(s) == 6:
                return f"{s[:4]}-{s[4:6]}-01"
        return str(val)
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        # Excel serial date (1900-01-01 = 1)
        if 1 <= val <= 100000:
            try:
                import datetime as _dt
                # Excel bug: treats 1900-02-29 as valid; use 1899-12-30 as base
                base = _dt.date(1899, 12, 30)
                delta = _dt.timedelta(days=int(val))
                return (base + delta).isoformat()
            except Exception:
                return str(val)
        return str(val)
    # datetime / date
    if hasattr(val, "isoformat"):
        return val.isoformat()[:10]
    return str(val)


def extract_criticas(sheets, cadoc_code: str) -> list:
    criticas = []
    for sheet in sheets:
        criticas.extend(extract_criticas_from_sheet(sheet["rows"], sheet["name"], cadoc_code))
    return criticas


def extract_leiautes(sheets, cadoc_code: str, source: str) -> dict:
    """Extrai campos preservando estrutura completa."""
    all_rows = []
    for sheet in sheets:
        for row in sheet["rows"]:
            if not row or not any(c is not None for c in row):
                continue
            # Remove None trailing
            while row and row[-1] is None:
                row.pop()
            if not row:
                continue
            record = {
                "sheet": sheet["name"],
                "values": [str(c).strip() if c is not None else "" for c in row],
            }
            all_rows.append(record)
    return {
        "source": source,
        "total_rows": len(all_rows),
        "rows": all_rows,
    }


def main():
    log("=" * 60)
    log("EXTRAÇÃO REFINADA v3 — Radiant Sentinel")
    log("=" * 60)

    criticas_total = {}
    leiautes_total = {}

    # === CRÍTICAS ===
    log("\n--- CRÍTICAS ---")
    criticas_files = [
        ("3040", "3040/SCR3040_Criticas.xls"),
        ("3050", "3050/Criticas_TXB_V11.xlsx"),  # V11 (Sprint 2 C.1) — mais recente que V9
        ("2061-DLO", "_referencias/Criticas_Processamento_2061_2071.xlsx"),
        ("2070-DDR", "2070-DDR/2070_DDR_Criticas.xlsx"),
    ]
    for cadoc, rel_path in criticas_files:
        p = ROOT / rel_path
        if not p.exists():
            log(f"  ⚠ {cadoc}: arquivo não encontrado")
            continue
        log(f"  → {cadoc}: {p.name}")
        sheets = read_any(p)
        criticas = extract_criticas(sheets, cadoc)
        if criticas:
            criticas_total[cadoc] = criticas
            log(f"    ✓ {len(criticas)} críticas extraídas")
            for c in criticas[:2]:
                log(f"      ex: {c.get('codigo')} — {c.get('regra', c.get('crítica', c.get('critica', '')))[:60]}")

    # Carrega DRM críticas (extraídas do PDF — Sprint 2 C.2)
    drm_path = OUT / "drm_criticas_raw.json"
    if drm_path.exists():
        log(f"  → 2060-DRM: {drm_path.name} (extraído do PDF em Sprint 2)")
        drm_data = json.loads(drm_path.read_text(encoding="utf-8"))
        criticas_total["2060-DRM"] = drm_data["criticas"]
        log(f"    ✓ {len(drm_data['criticas'])} críticas DRM adicionadas")

    # === LEIAUTES ===
    log("\n--- LEIAUTES ---")
    leiautes_files = [
        ("3040", "3040/SCR3040_Leiaute.xls"),
        ("3042", "3042/SCR3042_Leiaute.xls"),
        ("3050", "3050/Leiaute_TXB_XML_V3.xls"),
        ("2030-DRSAC", "2030-DRSAC/2030_DRSAC_Leiaute.xlsx"),
        ("2060-DRM", "2060-DRM/2060_DRM_Leiaute.xls"),
        ("2062-DLI", "2062-DLI/2062_DLI_Leiaute.xlsx"),
        ("2070-DDR", "2070-DDR/DDR_2011_Leiaute.xls"),
        ("2160-DRL", "2160-DRL/2160_DRL_Leiaute.xlsx"),
    ]
    for cadoc, rel_path in leiautes_files:
        p = ROOT / rel_path
        if not p.exists():
            log(f"  ⚠ {cadoc}: arquivo não encontrado")
            continue
        log(f"  → {cadoc}: {p.name}")
        sheets = read_any(p)
        leiaute = extract_leiautes(sheets, cadoc, p.name)
        if leiaute["total_rows"] > 0:
            leiautes_total[cadoc] = leiaute
            log(f"    ✓ {leiaute['total_rows']} linhas de campos")

    # === SALVAR ===
    log("\n--- SALVANDO ---")
    criticas_path = OUT / "criticas.json"
    leiautes_path = OUT / "leiautes.json"

    with open(criticas_path, "w", encoding="utf-8") as f:
        json.dump({
            "_metadata": {
                "description": "Catálogo consolidado de regras de validação (críticas) dos CADOCs BACEN",
                "fonte": "Planilhas oficiais BACEN (Deinf/Dine4 / áreas técnicas)",
                "data_extracao": datetime.now().isoformat(timespec="seconds"),
                "autor": "Mavis · Radiant Risk Solutions",
                "uso": "Sentinel Audit L2 (validação semântica) — base para reimplementação em Go",
                "total_cadocs": len(criticas_total),
                "total_criticas": sum(len(v) for v in criticas_total.values()),
            },
            "criticas": criticas_total,
        }, f, ensure_ascii=False, indent=2, default=str)
    log(f"  ✓ {criticas_path.name}: {criticas_path.stat().st_size:,} bytes")

    with open(leiautes_path, "w", encoding="utf-8") as f:
        json.dump({
            "_metadata": {
                "description": "Catálogo consolidado de leiautes (estrutura de campos) dos CADOCs BACEN",
                "fonte": "Planilhas oficiais BACEN",
                "data_extracao": datetime.now().isoformat(timespec="seconds"),
                "autor": "Mavis · Radiant Risk Solutions",
                "uso": "Schema Registry + XSD generator — base para L1 (validação estrutural)",
                "total_cadocs": len(leiautes_total),
                "total_rows": sum(v["total_rows"] for v in leiautes_total.values()),
            },
            "leiautes": leiautes_total,
        }, f, ensure_ascii=False, indent=2, default=str)
    log(f"  ✓ {leiautes_path.name}: {leiautes_path.stat().st_size:,} bytes")

    log("\n" + "=" * 60)
    total_criticas = sum(len(v) for v in criticas_total.values())
    total_leiautes_rows = sum(v["total_rows"] for v in leiautes_total.values())
    log(f"RESUMO:")
    log(f"  Críticas: {len(criticas_total)} CADOCs / {total_criticas} regras")
    log(f"  Leiautes: {len(leiautes_total)} CADOCs / {total_leiautes_rows} linhas")
    log("=" * 60)

    with open(OUT / "_extracao.log", "w", encoding="utf-8") as f:
        f.write("\n".join(LOG))


if __name__ == "__main__":
    main()